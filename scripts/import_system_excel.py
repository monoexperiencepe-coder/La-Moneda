# -*- coding: utf-8 -*-
"""Export SYSTEM (2).xlsx to JSON seed for La Moneda.

Enriquece INGRESOS/GASTOS con todas las columnas del Excel: campos dedicados,
heurísticas (método de pago, banco, N° operación, periodo, responsable) si existen
en el archivo, y ``excelColumnasExtra`` con el resto serializado.
Salida por defecto: ``src/data/systemExcelSeed_v2.json``.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    raise


def deref(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def num_id(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        x = float(v)
        if x != x:  # nan
            return None
        return int(x)
    except (TypeError, ValueError):
        return None


def num_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        x = float(v)
        if x != x:
            return None
        return x
    except (TypeError, ValueError):
        return None


def str_clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def norm_header(h) -> str:
    if h is None:
        return ""
    s = " ".join(str(h).strip().split()).upper()
    for a, b in (
        ("Á", "A"),
        ("É", "E"),
        ("Í", "I"),
        ("Ó", "O"),
        ("Ú", "U"),
        ("Ñ", "N"),
        ("°", ""),
        ("º", ""),
    ):
        s = s.replace(a, b)
    return s


def cell_to_json(v):
    """Valor serializable a JSON (fechas como ISO)."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, bool):
        return v
    if isinstance(v, (int,)):
        return v
    if isinstance(v, float):
        if v != v:
            return None
        if v == int(v):
            return int(v)
        return v
    return str_clean(v)


def parse_fecha_cell(v):
    """Return YYYY-MM-DD or None."""
    x = deref(v)
    if x is None:
        return None
    if isinstance(x, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(x[:10], fmt).date().isoformat()
            except ValueError:
                continue
        if re.match(r"^\d{4}-\d{2}-\d{2}", x):
            return x[:10]
    return None


def map_ingreso_tipo(raw: str) -> tuple[str, str | None]:
    t = (raw or "").strip().upper()
    t = t.replace("Í", "I").replace("Á", "A")  # normalize partial
    if "ALQUILER" in t or t == "ALQUILER":
        return "ALQUILER", "Día"
    if "GARANT" in t:
        return "GARANTÍAS", "PARCIAL"
    if "PAPELET" in t:
        return "OTROS INGRESOS", "PAPELETAS/MYLTAS"
    if "INTERES" in t:
        return "CAPITAL DE TRABAJO", "Préstamo personas"
    if "PRESTAM" in t:
        return "CAPITAL DE TRABAJO", "Préstamo personas"
    if "APORT" in t:
        return "APORTES", "Transferencia bancaria"
    if "DEVOLUC" in t and "PREST" in t:
        return "DEVOLUCION PRESTAMO", "Capital"
    if "CHOQUE" in t or "DAÑO" in t or "DANO" in t:
        return "OTROS INGRESOS", "CHOQUES / DAÑOS"
    return raw.strip() if raw else "OTROS INGRESOS", "Otras"


def gasto_tipo_from_row(tipo, categoria, comentarios) -> tuple[str, str]:
    t = str_clean(tipo)
    c = str_clean(categoria)
    if t and t.upper() not in ("NONE", "NULL"):
        return t, c or t
    if c and c.upper() not in ("NONE", "NULL"):
        return c, c
    com = str_clean(comentarios)[:120]
    if not com:
        return "OTROS GASTOS", "Sin detalle"
    return "MECÁNICOS", com[:80]


def is_reg_fechas_meta_column(h: str | None) -> bool:
    """Columnas de la hoja REG. FECHAS que no son vencimientos (ids, notas, fecha base)."""
    if not h:
        return True
    lt = str(h).strip()
    u = lt.upper()
    if u.startswith("FECHA") and "AUTO" not in u:
        return True
    if u in ("DETALLE AUTO", "NOTAS"):
        return True
    nk = norm_header(lt).replace(" ", "")
    if "AUTO" in nk and nk.startswith("N"):
        return True
    return False


def _reg_fecha_header_matched(h: str) -> str | None:
    """
    Mapea encabezado de columna de vencimiento → tipo ControlFecha (código app).
    Usa norm_header para tolerar acentos y símbolos. None = sin regla explícita
    (el import usará OTRO_VENCIMIENTO si la celda tiene fecha).
    """
    if not h:
        return None
    raw_u = str(h).strip().upper()
    s = norm_header(h)
    sc = s.replace(" ", "").replace(".", "").replace("-", "").replace("/", "")

    # Batería
    if "BATMANT" in sc or ("BAT" in s and "MANT" in s and "REALIZ" in s):
        return "BAT_MANT_REALIZADO"
    if ("BAT" in s and "COMPRA" in s) or "BATCOMPRA" in sc:
        return "BAT_COMPRA_NUEVA"

    # SOAT / seguro obligatorio
    if "SOAT" in s or ("SEGURO" in s and "OBLIGATORIO" in s):
        return "SOAT"

    # AFOCAT antes que reglas genéricas con "TAXI"
    if "AFOCAT" in s:
        return "AFOCAT_TAXI"

    # Credencial / brevete (antes que venc. brevete genérico)
    if "CREDENCIAL" in s and ("BREVETE" in s or "BREVET" in s):
        return "CREDENCIAL_ATU_BREVETE"
    if "VENC" in s and ("BREVETE" in s or "BREVET" in s):
        return "VENC_BREVETE"
    if "BREVETE" in s or "BREVET" in s:
        return "VENC_BREVETE"

    # Revisión técnica / RT (particular vs taxi / Detaxi)
    if "DETAXI" in s or "DETAXI" in sc or ("DET" in s and "TAXI" in s):
        return "RT_TAXI"
    has_part = "PARTICULAR" in s or "PARTICUL" in s
    has_rt = "RT" in s or "R.T" in raw_u
    has_revm = "REVISION" in s or "RTV" in s or "RVT" in sc
    if has_part and (has_rt or has_revm or "TECNIC" in s):
        return "RT_PARTICULAR"
    if ("TAXI" in s or "TAXI" in sc) and (has_rt or has_revm or "TECNIC" in s):
        return "RT_TAXI"
    if has_revm and "TECNIC" in s:
        return "RT_TAXI"
    if has_rt:
        if has_part:
            return "RT_PARTICULAR"
        if "TAXI" in s:
            return "RT_TAXI"
    if "INSTALACION" in s and "GNV" in s:
        return "INSTALACION_GNV"
    if "PERMISO" in s and "ATU" in s:
        return "PERMISO_ATU"
    if ("CERTIFICADO" in s or "CERTIF" in s) and ("ANUAL" in s or "GNV" in s):
        return "CERT_GNV_ANUAL"
    if "QUINQUENAL" in s or "QUINQUEN" in s:
        return "QUINQUENAL_GNV"

    # GPS / rastreo
    if "GPS" in s or "RASTREO" in s or "TRACKING" in s or ("PLATAFORMA" in s and "GPS" in s):
        return "GPS"

    # Impuesto vehicular / tributo
    if "IMPUEST" in s or ("TRIBUTO" in s and "VEHIC" in s) or ("SUNARP" in s and "VEHIC" in s):
        return "IMPUESTO"
    if "VENC" in s and "IMPUEST" in s:
        return "IMPUESTO"

    return None


def reg_fecha_header_to_tipo(h: str) -> str | None:
    """Compat: regla explícita o None (el bucle REG. FECHAS aplica OTRO_VENCIMIENTO)."""
    return _reg_fecha_header_matched(h)


def audit_reg_fechas_headers(ws) -> dict:
    """Diagnóstico: cabeceras meta, mapeadas a tipo, y sin regla explícita (→ OTRO si hay fecha)."""
    hdrs = sheet_headers(ws)
    meta = []
    mapeadas: dict[str, str] = {}
    sin_regla: list[str] = []
    for h in hdrs:
        if is_reg_fechas_meta_column(h):
            meta.append(h)
            continue
        t = _reg_fecha_header_matched(h)
        if t:
            mapeadas[str(h)] = t
        else:
            sin_regla.append(str(h))
    return {"meta_columnas": meta, "mapeadas": mapeadas, "sin_regla_explicita": sin_regla}


def iter_dict_rows(ws, min_row=1):
    rows = ws.iter_rows(min_row=min_row, values_only=True)
    headers = next(rows)
    hdrs = [h if h is not None else f"__{i}" for i, h in enumerate(headers)]
    for r in rows:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        yield dict(zip(hdrs, r))


def sheet_headers(ws, min_row=1) -> list[str]:
    row = next(ws.iter_rows(min_row=min_row, max_row=min_row, values_only=True), None)
    if not row:
        return []
    return [str(h) if h is not None else f"__{i}" for i, h in enumerate(row)]


def _pick_first_matching_row_value(row: dict, rules: list[tuple[str, ...]]) -> tuple[str | None, object | None]:
    """Primera columna cuyo encabezado cumple todas las subcadenas de una regla."""
    for k, v in row.items():
        if k is None or v in (None, ""):
            continue
        nk = norm_header(k)
        for subs in rules:
            if all(s in nk for s in subs):
                return str(k), v
    return None, None


def heuristic_finance_columns(row: dict) -> tuple[dict, set[str]]:
    """Mapea columnas típicas si existen (archivos ampliados o otras versiones)."""
    out = {
        "metodoPagoReal": None,
        "bancoCanal": None,
        "numeroOperacion": None,
        "fechaPeriodoDesde": None,
        "fechaPeriodoHasta": None,
        "usuarioResponsable": None,
    }
    used: set[str] = set()
    mk, mv = _pick_first_matching_row_value(
        row,
        [
            ("METODO", "PAGO"),
            ("MEDIO", "PAGO"),
            ("FORMA", "PAGO"),
            ("MEDIO", "COBRO"),
            ("TIPO", "PAGO"),
        ],
    )
    if mk:
        used.add(mk)
        out["metodoPagoReal"] = str_clean(mv) or None
    bk, bv = _pick_first_matching_row_value(
        row,
        [
            ("BANCO",),
            ("ENTIDAD",),
            ("CANAL", "PAGO"),
            ("CUENTA", "DESTINO"),
        ],
    )
    if bk:
        used.add(bk)
        out["bancoCanal"] = str_clean(bv) or None
    ok, ov = _pick_first_matching_row_value(
        row,
        [
            ("OPERACION",),
            ("VOUCHER",),
            ("TRANSACCION",),
            ("NUM", "OPER"),
            ("NRO", "OPER"),
            ("COD", "OPER"),
        ],
    )
    if ok:
        used.add(ok)
        out["numeroOperacion"] = str_clean(ov) or None
    dk, dv = _pick_first_matching_row_value(
        row,
        [
            ("VIGENCIA", "DESDE"),
            ("PERIODO", "DESDE"),
            ("FECHA", "DESDE"),
        ],
    )
    if dk:
        used.add(dk)
        out["fechaPeriodoDesde"] = parse_fecha_cell(dv) or (
            dv.date().isoformat() if isinstance(dv, datetime) else None
        )
    hk, hv = _pick_first_matching_row_value(
        row,
        [
            ("VIGENCIA", "HASTA"),
            ("PERIODO", "HASTA"),
            ("FECHA", "HASTA"),
        ],
    )
    if hk:
        used.add(hk)
        out["fechaPeriodoHasta"] = parse_fecha_cell(hv) or (
            hv.date().isoformat() if isinstance(hv, datetime) else None
        )
    uk, uv = _pick_first_matching_row_value(
        row,
        [
            ("USUARIO",),
            ("RESPONSABLE",),
            ("REGISTRADO", "POR"),
            ("AUTOR",),
            ("OPERADOR",),
        ],
    )
    if uk:
        used.add(uk)
        out["usuarioResponsable"] = str_clean(uv) or None
    return out, used


def periodo_desde_hasta_excel(anio: int | None, mes: int | None) -> tuple[str | None, str | None]:
    """Primer y último día del mes si hay año y mes válidos en el Excel."""
    if anio is None or mes is None:
        return None, None
    try:
        y, m = int(anio), int(mes)
        if y < 1900 or m < 1 or m > 12:
            return None, None
        d1 = date(y, m, 1)
        d2 = date(y, m, monthrange(y, m)[1])
        return d1.isoformat(), d2.isoformat()
    except (TypeError, ValueError, OverflowError):
        return None, None


def excel_columnas_extra(row: dict, used_keys: set) -> dict[str, object]:
    extra = {}
    for k, v in row.items():
        if k is None or k in used_keys:
            continue
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        key = str(k).strip()
        if not key or key.startswith("__"):
            continue
        j = cell_to_json(v)
        if j is not None:
            extra[key] = j
    return extra


def find_ncarro_column_key(row: dict) -> str | None:
    for k in row:
        if k is None:
            continue
        n = norm_header(k).replace(" ", "")
        if "CARRO" in n and n.startswith("N"):
            return str(k)
    return None


def find_detalle_auto_key(row: dict) -> str | None:
    for k in row:
        if k is None:
            continue
        nk = norm_header(k)
        if "DETALLE" in nk and "AUTO" in nk:
            return str(k)
    return None


def find_categoria_key(row: dict) -> str | None:
    for k in row:
        if k and "CATEGOR" in str(k).upper():
            return str(k)
    return None


def join_detalle_operativo(detalle_auto: str, comentarios: str) -> str | None:
    a = str_clean(detalle_auto)
    c = str_clean(comentarios)
    if a and c:
        return f"{a} — {c}"
    return a or c or None


def infer_estado_pago_comentarios(comentarios: str) -> str | None:
    u = str_clean(comentarios).upper()
    if not u:
        return None
    for needle in ("PENDIENTE", "DEBE", "POR COBRAR", "POR PAGAR"):
        if needle in u:
            return "PENDIENTE"
    for needle in ("PAGADO", "COBRADO", "LIQUIDADO"):
        if needle in u:
            return "PAGADO"
    return None


def tipo_operacion_ingreso(tipo: str, sub: str | None) -> str | None:
    parts = [str_clean(tipo)]
    if sub:
        parts.append(str_clean(sub))
    parts = [p for p in parts if p]
    return " | ".join(parts) if parts else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "xlsx",
        nargs="?",
        default=str(Path.home() / "Downloads" / "SYSTEM (2).xlsx"),
        help="Path to SYSTEM (2).xlsx",
    )
    ap.add_argument(
        "-o",
        "--out",
        default=str(Path(__file__).resolve().parents[1] / "src" / "data" / "systemExcelSeed_v2.json"),
    )
    args = ap.parse_args()
    src = Path(args.xlsx)
    if not src.is_file():
        print(f"Missing file: {src}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(src, data_only=True, read_only=False)
    ing_headers = sheet_headers(wb["INGRESOS"])
    gas_headers = sheet_headers(wb["GASTOS"])

    reg_audit = audit_reg_fechas_headers(wb["REG. FECHAS"])
    print(
        "REG. FECHAS — auditoría de cabeceras:",
        json.dumps(reg_audit, ensure_ascii=False, indent=2),
        file=sys.stderr,
    )

    vehicles = []
    unidades = []
    vid = 9000
    uid = 10000
    for row in iter_dict_rows(wb["UNIDADES"]):
        n_raw = row.get("N°")
        if n_raw is None:
            n_raw = row.get("N�")  # mojibake fallback
        nid = num_id(n_raw)
        if nid is None:
            continue
        marca = str_clean(row.get("MARCA"))
        modelo = str_clean(row.get("MODELO"))
        anio = num_id(row.get("AÑO"))
        if anio is None:
            anio = num_id(row.get("A�O"))
        placa = str_clean(row.get("PLACA"))
        color = str_clean(row.get("COLOR"))
        vehicles.append(
            {
                "id": nid,
                "marca": marca or "—",
                "modelo": modelo or "—",
                "placa": placa or "—",
                "anio": anio if anio else None,
                "color": color or None,
                "activo": True,
            }
        )
        kmi = num_id(row.get("KM INICIAL"))
        llaves = num_id(row.get("# LLAVES"))
        vc = num_float(row.get("V. COMPRA USD"))
        tc = row.get("TIPO CAMBIO")
        tc_num = num_float(tc)
        fc = parse_fecha_cell(row.get("F. COMPRA USD"))
        gnv = row.get("GASTO GNV")
        gn = row.get("GASTOS NOTARIALES")
        ga = row.get("GASTOS ACCSESORIOS")
        extra_com = ""
        if tc_num is None and str_clean(tc) and str_clean(tc) not in ("?", "-", "—"):
            extra_com = f"TC compra (texto): {str_clean(tc)}"
        gps1 = str_clean(row.get("# GPS 1")) or None
        gps2 = str_clean(row.get("# GPS 2")) or None
        unidades.append(
            {
                "id": uid,
                "vehicleId": nid,
                "numeroInterno": str(nid),
                "marca": marca or "—",
                "modelo": modelo or "—",
                "anio": int(anio) if anio else 0,
                "placa": placa or "—",
                "detalleAuto": str_clean(row.get("DETALLE DEL AUTO")),
                "combustible": str_clean(row.get("COMBUSTIBLE")),
                "color": color,
                "tipoCarroceria": str_clean(row.get("TIPO CARROCERÍA")) or str_clean(row.get("TIPO CARROCER�A")),
                "numeroMotor": str_clean(row.get("N MOTOR")) or None,
                "cantidadLlaves": llaves,
                "gps1": gps1,
                "gps2": gps2,
                "impuestoEstado": str_clean(row.get("IMPUESTO")) or None,
                "kmInicial": kmi,
                "tarjetaPropiedad": str_clean(row.get("TARJ. PROPIEDAD")) or None,
                "propietario": str_clean(row.get("PROPIETARIO")) or None,
                "fechaCompraUSD": fc,
                "valorCompraUSD": vc,
                "tipoCambioCompra": tc_num,
                "gastoGnv": str_clean(gnv) if gnv not in (None, "") else None,
                "gastosNotariales": str_clean(gn) if gn not in (None, "") else None,
                "gastosAccesorios": str_clean(ga) if ga not in (None, "") else None,
                "gpsInstalado": bool(gps1 or gps2),
                "gpsProveedor": "GPS FLOTA",
                "impuestoVehicularVence": None,
                "comentarios": extra_com,
                "createdAt": "2025-01-01T10:00:00",
            }
        )
        uid += 1

    vehicles.sort(key=lambda x: x["id"])
    # dedupe vehicle ids (keep first)
    seen_v = set()
    vehicles = [v for v in vehicles if v["id"] not in seen_v and not seen_v.add(v["id"])]

    conductores_out = []
    cid = 20000
    for row in iter_dict_rows(wb["CONDUCTORES"]):
        car = num_id(row.get("CARRO ASIGNADO"))
        if car is None:
            continue
        doc_t = str_clean(row.get("DOCUMENTO")).upper()
        if doc_t not in ("DNI", "CE", "PASAPORTE"):
            doc_t = "DNI"
        dom = str_clean(row.get("DOMICILIO")).upper()
        if "PROPIO" in dom:
            domicilio = "PROPIO"
        elif "FAMILIA" in dom or "CASA DE" in dom:
            domicilio = "CASA DE FAMILIA"
        else:
            domicilio = "ALQUILADO"
        st_raw = str_clean(row.get("STATUS"))
        st = st_raw.upper()
        estado = "VIGENTE" if "VIGEN" in st or st == "" or "ACTIV" in st else "SUSPENDIDO"
        coch = str_clean(row.get("COCHERA"))
        freg = parse_fecha_cell(row.get("FECHA REGISTRO")) or "2024-01-01"
        fv = parse_fecha_cell(row.get("F.V CONTRATO"))
        df = row.get("DOC. FIRMADO")
        doc_firm = None
        if isinstance(df, bool):
            doc_firm = df
        elif str_clean(df).lower() in ("si", "sí", "1", "true", "x"):
            doc_firm = True
        elif str_clean(df).lower() in ("no", "0", "false"):
            doc_firm = False
        conductores_out.append(
            {
                "id": cid,
                "vehicleId": car,
                "tipoDocumento": doc_t,
                "numeroDocumento": str_clean(row.get("NÚMERO")) or str_clean(row.get("N�MERO")),
                "nombres": str_clean(row.get("NOMBRES")),
                "apellidos": str_clean(row.get("APELLIDOS")),
                "celular": str_clean(row.get("CELULAR")),
                "domicilio": domicilio,
                "estadoContrato": "CERRADO",
                "estado": estado,
                "statusOriginal": st_raw or None,
                "cochera": coch or None,
                "numeroEmergencia": str_clean(row.get("NUMERO DE EMERGENCIA")) or None,
                "direccion": str_clean(row.get("DIRECCIÓN")) or str_clean(row.get("DIRECCI�N")) or None,
                "documentoFirmado": doc_firm,
                "fechaVencimientoContrato": fv,
                "comentarios": "",
                "createdAt": freg + "T09:00:00",
            }
        )
        cid += 1

    control_rows = []
    ctid = 30000
    for row in iter_dict_rows(wb["REG. FECHAS"]):
        car = num_id(row.get("N° AUTO"))
        if car is None:
            car = num_id(row.get("N� AUTO"))
        if car is None:
            continue
        base_fecha = parse_fecha_cell(row.get("FECHA")) or date.today().isoformat()
        notas = str_clean(row.get("NOTAS"))
        for hk, val in row.items():
            if hk is None or val is None or str(val).strip() == "":
                continue
            lt = str(hk).strip()
            if is_reg_fechas_meta_column(lt):
                continue
            fv = parse_fecha_cell(val)
            if not fv:
                continue
            tipo = _reg_fecha_header_matched(lt)
            if not tipo:
                tipo = "OTRO_VENCIMIENTO"
                col_note = f"[Columna Excel: {lt}]"
                com_ctrl = f"{col_note} {notas}".strip() if notas else col_note
            else:
                com_ctrl = notas
            control_rows.append(
                {
                    "id": ctid,
                    "vehicleId": car,
                    "tipo": tipo,
                    "fechaVencimiento": fv,
                    "fechaRegistro": base_fecha,
                    "comentarios": com_ctrl,
                    "createdAt": base_fecha + "T08:00:00",
                }
            )
            ctid += 1

    km_out = []
    kmid = 40000
    for row in iter_dict_rows(wb["KMS"]):
        car = num_id(row.get("N° AUTO"))
        if car is None:
            car = num_id(row.get("N� AUTO"))
        if car is None:
            continue
        dia_val = None
        for k in row:
            if not k:
                continue
            ks = str(k).strip().upper().replace("Í", "I")
            if ks in ("DIA", "DAY"):
                dia_val = row.get(k)
                break
        if dia_val is None:
            dia_val = row.get("Día") or row.get("DÍA")
        fd = parse_fecha_cell(dia_val)
        freg_cell = row.get("Fecha de registro")
        freg = parse_fecha_cell(freg_cell)
        if not freg and isinstance(freg_cell, datetime):
            freg = freg_cell.date().isoformat()
        if not freg:
            freg = fd or date.today().isoformat()
        if not fd:
            fd = freg
        km_m = num_float(row.get("KM MANT."))
        km = num_float(row.get("KILOMETRAJE"))
        desc = str_clean(row.get("DESCRIPCIÓN")) or str_clean(row.get("DESCRIPCI�N"))
        cost = num_float(row.get("COSTO"))
        km_out.append(
            {
                "id": kmid,
                "vehicleId": car,
                "fecha": fd,
                "fechaRegistro": freg,
                "kmMantenimiento": int(km_m) if km_m is not None else None,
                "kilometraje": int(km) if km is not None else None,
                "descripcion": desc,
                "costo": cost,
                "createdAt": freg + "T09:00:00",
            }
        )
        kmid += 1

    ing_out = []
    iid = 50000
    for row in iter_dict_rows(wb["INGRESOS"]):
        used_keys: set[str] = set()

        ck = find_ncarro_column_key(row)
        if not ck:
            continue
        used_keys.add(ck)
        car = num_id(row[ck])
        if car is None:
            continue
        used_keys.add("MONTO")
        monto = num_float(row.get("MONTO"))
        if monto is None:
            continue
        used_keys.add("TIPO")
        raw_tipo = str_clean(row.get("TIPO"))
        tipo, sub = map_ingreso_tipo(raw_tipo)
        used_keys.add("FECHA")
        fecha = parse_fecha_cell(row.get("FECHA"))
        if not fecha:
            continue
        freg_cell = row.get("F.REGISTRO")
        if "F.REGISTRO" in row:
            used_keys.add("F.REGISTRO")
        freg = parse_fecha_cell(freg_cell)
        if not freg and isinstance(freg_cell, datetime):
            freg = freg_cell.date().isoformat()
        elif not freg:
            freg = fecha
        fr_dt = row.get("F.REGISTRO")
        created = fecha + "T08:00:00"
        if isinstance(fr_dt, datetime):
            created = fr_dt.isoformat(timespec="seconds")
        used_keys.add("COMENTARIOS")
        com = str_clean(row.get("COMENTARIOS"))

        detalle = ""
        dk = find_detalle_auto_key(row)
        if dk:
            used_keys.add(dk)
            detalle = str_clean(row[dk])

        anio_ex = None
        mes_ex = None
        for ak in ("AÑO", "AO", "ANO"):
            if ak in row and row[ak] not in (None, ""):
                used_keys.add(ak)
                anio_ex = num_id(row[ak])
                break
        if "MES" in row and row["MES"] not in (None, ""):
            used_keys.add("MES")
            mes_ex = num_id(row["MES"])

        heur, hused = heuristic_finance_columns(row)
        used_keys |= hused

        p_desde_m, p_hasta_m = periodo_desde_hasta_excel(anio_ex, mes_ex)
        f_desde = heur["fechaPeriodoDesde"] or p_desde_m
        f_hasta = heur["fechaPeriodoHasta"] or p_hasta_m

        extra = excel_columnas_extra(row, used_keys)

        det_op = join_detalle_operativo(detalle, com)
        tipo_op = tipo_operacion_ingreso(tipo, sub)
        est_pago = infer_estado_pago_comentarios(com)

        ing_out.append(
            {
                "id": iid,
                "fecha": fecha,
                "fechaRegistro": freg,
                "vehicleId": car,
                "tipo": tipo,
                "subTipo": sub,
                "fechaDesde": f_desde,
                "fechaHasta": f_hasta,
                "metodoPago": "Otro",
                "metodoPagoDetalle": "Excel",
                "metodoPagoReal": heur["metodoPagoReal"],
                "bancoCanal": heur["bancoCanal"],
                "numeroOperacion": heur["numeroOperacion"],
                "fechaPeriodoDesde": heur["fechaPeriodoDesde"],
                "fechaPeriodoHasta": heur["fechaPeriodoHasta"],
                "usuarioResponsable": heur["usuarioResponsable"],
                "detalleDelAuto": detalle or None,
                "periodoAnioExcel": anio_ex,
                "periodoMesExcel": mes_ex,
                "detalleOperativo": det_op,
                "tipoOperacion": tipo_op,
                "estadoPago": est_pago,
                "celularMetodo": None,
                "signo": "+",
                "monto": float(monto),
                "moneda": "PEN",
                "tipoCambio": None,
                "montoPENReferencia": float(monto),
                "comentarios": com,
                "excelColumnasExtra": extra,
                "createdAt": created,
            }
        )
        iid += 1

    g_out = []
    gid = 60000
    for row in iter_dict_rows(wb["GASTOS"]):
        used_keys: set[str] = set()
        used_keys.add("MONTO")
        monto_raw = num_float(row.get("MONTO"))
        if monto_raw is None:
            continue
        monto = abs(monto_raw)

        ck = find_ncarro_column_key(row)
        if not ck:
            continue
        used_keys.add(ck)
        car = num_id(row[ck])

        used_keys.add("FECHA")
        fecha = parse_fecha_cell(row.get("FECHA"))
        if not fecha:
            continue
        freg_cell = row.get("F.REGISTRO")
        if "F.REGISTRO" in row:
            used_keys.add("F.REGISTRO")
        freg = parse_fecha_cell(freg_cell)
        if not freg and isinstance(freg_cell, datetime):
            freg = freg_cell.date().isoformat()
        elif not freg:
            freg = fecha
        fr_dt = row.get("F.REGISTRO")
        created = fecha + "T10:00:00"
        if isinstance(fr_dt, datetime):
            created = fr_dt.isoformat(timespec="seconds")

        cat = None
        cat_key = find_categoria_key(row)
        if cat_key:
            used_keys.add(cat_key)
            cat = row.get(cat_key)
        categoria_excel_raw = str_clean(cat) if cat is not None else None

        used_keys.add("TIPO")
        used_keys.add("COMENTARIOS")
        tipo, sub = gasto_tipo_from_row(row.get("TIPO"), cat, row.get("COMENTARIOS"))
        com = str_clean(row.get("COMENTARIOS"))

        detalle = ""
        dk = find_detalle_auto_key(row)
        if dk:
            used_keys.add(dk)
            detalle = str_clean(row[dk])

        anio_ex = None
        mes_ex = None
        for ak in ("AÑO", "AO", "ANO"):
            if ak in row and row[ak] not in (None, ""):
                used_keys.add(ak)
                anio_ex = num_id(row[ak])
                break
        if "MES" in row and row["MES"] not in (None, ""):
            used_keys.add("MES")
            mes_ex = num_id(row["MES"])

        heur, hused = heuristic_finance_columns(row)
        used_keys |= hused

        p_desde_m, p_hasta_m = periodo_desde_hasta_excel(anio_ex, mes_ex)
        f_desde = heur["fechaPeriodoDesde"] or p_desde_m
        f_hasta = heur["fechaPeriodoHasta"] or p_hasta_m

        extra = excel_columnas_extra(row, used_keys)

        g_det_op = join_detalle_operativo(detalle, com)
        cat_excel = str_clean(cat) if cat is not None else ""
        categoria_real = cat_excel or str_clean(tipo) or None
        subcat = str_clean(sub) or None

        g_out.append(
            {
                "id": gid,
                "fecha": fecha,
                "fechaRegistro": freg,
                "vehicleId": car,
                "tipo": tipo,
                "subTipo": sub,
                "fechaDesde": f_desde,
                "fechaHasta": f_hasta,
                "motivo": sub,
                "metodoPago": "Otro",
                "metodoPagoDetalle": "Excel",
                "metodoPagoReal": heur["metodoPagoReal"],
                "bancoCanal": heur["bancoCanal"],
                "numeroOperacion": heur["numeroOperacion"],
                "fechaPeriodoDesde": heur["fechaPeriodoDesde"],
                "fechaPeriodoHasta": heur["fechaPeriodoHasta"],
                "usuarioResponsable": heur["usuarioResponsable"],
                "detalleOperativo": g_det_op,
                "categoriaReal": categoria_real,
                "subcategoria": subcat,
                "detalleDelAuto": detalle or None,
                "periodoAnioExcel": anio_ex,
                "periodoMesExcel": mes_ex,
                "categoriaExcelRaw": categoria_excel_raw,
                "celularMetodo": None,
                "categoria": "GASTOS_MECANICOS",
                "signo": "-",
                "monto": float(monto),
                "pagadoA": "",
                "comentarios": com,
                "excelColumnasExtra": extra,
                "createdAt": created,
            }
        )
        gid += 1

    wb.close()

    # Fix categorías gasto con reglas simples (alineado a inferCategoriaFromTipoGasto)
    def fix_gasto_cat(tipo: str) -> str:
        t = tipo.upper()
        if any(x in t for x in ("GASTOS FIJOS", "FIJO", "SUELDO", "SEGURO")):
            return "GASTOS_FIJOS"
        if any(x in t for x in ("DOCUMENT", "TRIBUT", "SOAT", "RT-", "AFOCAT", "NOTARIAL")):
            return "GASTOS_TRIBUTARIOS"
        if "OTROS GASTOS" in t:
            return "GASTOS_PROVISIONALES"
        return "GASTOS_MECANICOS"

    for g in g_out:
        g["categoria"] = fix_gasto_cat(g["tipo"])

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "source": str(src),
            "vehicles": len(vehicles),
            "unidades": len(unidades),
            "conductores": len(conductores_out),
            "controlFechas": len(control_rows),
            "kilometrajes": len(km_out),
            "ingresos": len(ing_out),
            "gastos": len(g_out),
            "ingresosColumnasExcel": ing_headers,
            "gastosColumnasExcel": gas_headers,
        },
        "vehicles": vehicles,
        "unidades": unidades,
        "conductores": conductores_out,
        "controlFechas": control_rows,
        "kilometrajes": km_out,
        "ingresos": ing_out,
        "gastos": g_out,
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print("Wrote", out_path, json.dumps(payload["meta"], ensure_ascii=False))


if __name__ == "__main__":
    main()
